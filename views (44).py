"""
Report Celery tasks: export generation, scheduling, and cleanup.
"""

import io
import logging
import time

from celery import shared_task
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def generate_report_export(self, export_id: str):
    """
    Generate a report export file (PDF, Excel, CSV, HTML).
    Called asynchronously via Celery.
    """
    from .models import ReportExport

    try:
        export = ReportExport.objects.select_related(
            "report", "report__organization"
        ).get(id=export_id)
    except ReportExport.DoesNotExist:
        logger.error("ReportExport %s not found.", export_id)
        return

    export.status = "processing"
    export.save(update_fields=["status"])

    start_time = time.time()

    try:
        report = export.report

        if export.format == "pdf":
            content, content_type = _generate_pdf(report, export.parameters_used)
            extension = "pdf"
        elif export.format == "excel":
            content, content_type = _generate_excel(report, export.parameters_used)
            extension = "xlsx"
        elif export.format == "csv":
            content, content_type = _generate_csv(report, export.parameters_used)
            extension = "csv"
        elif export.format == "html":
            content, content_type = _generate_html(report, export.parameters_used)
            extension = "html"
        else:
            content, content_type = _generate_markdown(report, export.parameters_used)
            extension = "md"

        filename = f"{report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{extension}"

        export.file.save(filename, ContentFile(content), save=False)
        export.file_size_bytes = len(content)
        export.status = "completed"
        export.generation_time_ms = int((time.time() - start_time) * 1000)
        export.save(update_fields=[
            "file", "file_size_bytes", "status", "generation_time_ms",
        ])

        logger.info(
            "Report export %s generated successfully in %dms.",
            export_id, export.generation_time_ms,
        )

    except Exception as exc:
        export.status = "failed"
        export.error_message = str(exc)
        export.save(update_fields=["status", "error_message"])
        logger.exception("Report export %s failed.", export_id)
        raise self.retry(exc=exc)


def _generate_pdf(report, parameters):
    """Generate a PDF report using ReportLab."""
    from reportlab.lib.pagesizes import A4, LETTER, LEGAL
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    page_sizes = {"a4": A4, "letter": LETTER, "legal": LEGAL}
    page_size = page_sizes.get(report.page_size, A4)

    if report.page_orientation == "landscape":
        page_size = page_size[1], page_size[0]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=page_size,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=24, spaceAfter=30,
    )
    heading_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading2"],
        fontSize=16, spaceBefore=20, spaceAfter=10,
    )
    body_style = styles["Normal"]

    elements = []

    if report.cover_page:
        elements.append(Spacer(1, 2 * inch))
        elements.append(Paragraph(report.title, title_style))
        elements.append(Paragraph(report.description or "", body_style))
        elements.append(Spacer(1, 0.5 * inch))
        elements.append(Paragraph(
            f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}",
            body_style,
        ))
        elements.append(Spacer(1, 2 * inch))

    for section in report.sections:
        section_title = section.get("title", "Untitled Section")
        section_type = section.get("type", "text")

        elements.append(Paragraph(section_title, heading_style))

        if section_type == "text":
            content = section.get("config", {}).get("content", "")
            elements.append(Paragraph(content, body_style))
        elif section_type == "metric":
            value = section.get("config", {}).get("value", "N/A")
            label = section.get("config", {}).get("label", "")
            elements.append(Paragraph(f"<b>{label}:</b> {value}", body_style))
        elif section_type == "table":
            table_data = section.get("config", {}).get("data", [])
            if table_data:
                headers = list(table_data[0].keys())
                data = [headers] + [[str(row.get(h, "")) for h in headers] for row in table_data]
                table = Table(data)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ecf0f1")]),
                ]))
                elements.append(table)

        elements.append(Spacer(1, 0.3 * inch))

    doc.build(elements)
    return buffer.getvalue(), "application/pdf"


def _generate_excel(report, parameters):
    """Generate an Excel report using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = report.title
    summary_ws["A1"].font = Font(size=16, bold=True)
    summary_ws["A2"] = report.description
    summary_ws["A3"] = f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}"

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for i, section in enumerate(report.sections):
        section_type = section.get("type", "text")
        section_title = section.get("title", f"Section {i + 1}")

        if section_type == "table":
            ws = wb.create_sheet(title=section_title[:31])
            table_data = section.get("config", {}).get("data", [])
            if table_data:
                headers = list(table_data[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                for row_idx, row in enumerate(table_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

                for col in ws.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

        elif section_type == "metric":
            row_num = summary_ws.max_row + 2
            label = section.get("config", {}).get("label", section_title)
            value = section.get("config", {}).get("value", "N/A")
            summary_ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_num, column=2, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _generate_csv(report, parameters):
    """Generate a CSV report from table sections."""
    import csv

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    for section in report.sections:
        if section.get("type") == "table":
            table_data = section.get("config", {}).get("data", [])
            if table_data:
                headers = list(table_data[0].keys())
                writer.writerow(headers)
                for row in table_data:
                    writer.writerow([row.get(h, "") for h in headers])
                writer.writerow([])

    return buffer.getvalue().encode("utf-8"), "text/csv"


def _generate_html(report, parameters):
    """Generate an HTML report."""
    sections_html = []
    for section in report.sections:
        section_type = section.get("type", "text")
        title = section.get("title", "")

        if section_type == "text":
            content = section.get("config", {}).get("content", "")
            sections_html.append(f"<h2>{title}</h2><p>{content}</p>")
        elif section_type == "metric":
            label = section.get("config", {}).get("label", title)
            value = section.get("config", {}).get("value", "N/A")
            change = section.get("config", {}).get("change", "")
            sections_html.append(
                f'<div class="metric-card"><h3>{label}</h3>'
                f'<span class="value">{value}</span>'
                f'<span class="change">{change}</span></div>'
            )
        elif section_type == "table":
            table_data = section.get("config", {}).get("data", [])
            if table_data:
                headers = list(table_data[0].keys())
                header_html = "".join(f"<th>{h}</th>" for h in headers)
                rows_html = ""
                for row in table_data:
                    cells = "".join(f"<td>{row.get(h, '')}</td>" for h in headers)
                    rows_html += f"<tr>{cells}</tr>"
                sections_html.append(
                    f"<h2>{title}</h2><table><thead><tr>{header_html}"
                    f"</tr></thead><tbody>{rows_html}</tbody></table>"
                )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{report.title}</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 40px; color: #2c3e50; }}
        h1 {{ color: #1a73e8; border-bottom: 2px solid #1a73e8; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
        th {{ background-color: #2c3e50; color: white; padding: 10px; text-align: left; }}
        td {{ padding: 8px 10px; border-bottom: 1px solid #ecf0f1; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        .metric-card {{ display: inline-block; background: #f8f9fa; border-radius: 8px; padding: 20px; margin: 10px; min-width: 200px; }}
        .metric-card .value {{ font-size: 2em; font-weight: bold; display: block; color: #1a73e8; }}
        .metric-card .change {{ font-size: 0.9em; color: #27ae60; }}
        .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #ecf0f1; color: #7f8c8d; font-size: 0.85em; }}
    </style>
</head>
<body>
    <h1>{report.title}</h1>
    <p>{report.description}</p>
    <p><em>Generated: {timezone.now().strftime('%B %d, %Y at %H:%M')}</em></p>
    {''.join(sections_html)}
    <div class="footer">
        <p>Report generated by InsightBoard</p>
    </div>
</body>
</html>"""
    return html.encode("utf-8"), "text/html"


def _generate_markdown(report, parameters):
    """Generate a Markdown report."""
    lines = [
        f"# {report.title}\n",
        f"{report.description}\n" if report.description else "",
        f"*Generated: {timezone.now().strftime('%B %d, %Y at %H:%M')}*\n\n---\n",
    ]

    for section in report.sections:
        section_type = section.get("type", "text")
        title = section.get("title", "")

        lines.append(f"\n## {title}\n")

        if section_type == "text":
            content = section.get("config", {}).get("content", "")
            lines.append(f"{content}\n")
        elif section_type == "metric":
            label = section.get("config", {}).get("label", title)
            value = section.get("config", {}).get("value", "N/A")
            lines.append(f"**{label}:** {value}\n")
        elif section_type == "table":
            table_data = section.get("config", {}).get("data", [])
            if table_data:
                headers = list(table_data[0].keys())
                lines.append("| " + " | ".join(headers) + " |")
                lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for row in table_data:
                    cells = " | ".join(str(row.get(h, "")) for h in headers)
                    lines.append(f"| {cells} |")
                lines.append("")

    return "\n".join(lines).encode("utf-8"), "text/markdown"


@shared_task
def process_scheduled_reports():
    """
    Check for reports that need to be generated based on their schedules.
    Runs every 15 minutes via Celery Beat.
    """
    from .models import ReportSchedule, ReportExport

    now = timezone.now()
    due_schedules = ReportSchedule.objects.filter(
        is_active=True,
        next_run_at__lte=now,
    ).select_related("report")

    for schedule in due_schedules:
        try:
            export = ReportExport.objects.create(
                report=schedule.report,
                schedule=schedule,
                format=schedule.export_format,
                generated_by=schedule.created_by,
                status="pending",
            )

            generate_report_export.delay(str(export.id))

            schedule.last_run_at = now
            _calculate_next_run(schedule)
            schedule.save(update_fields=["last_run_at", "next_run_at"])

            logger.info(
                "Triggered scheduled export for report '%s' (schedule %s).",
                schedule.report.title, schedule.id,
            )

        except Exception:
            logger.exception(
                "Failed to process schedule %s for report '%s'.",
                schedule.id, schedule.report.title,
            )


def _calculate_next_run(schedule):
    """Calculate the next run time for a schedule."""
    from datetime import timedelta
    from django.utils import timezone as tz

    now = tz.now()

    frequency_deltas = {
        "daily": timedelta(days=1),
        "weekly": timedelta(weeks=1),
        "biweekly": timedelta(weeks=2),
        "monthly": timedelta(days=30),
        "quarterly": timedelta(days=90),
    }

    delta = frequency_deltas.get(schedule.frequency, timedelta(days=1))
    schedule.next_run_at = now + delta


@shared_task
def cleanup_expired_exports():
    """
    Delete expired report exports and their files.
    Runs daily at 3 AM via Celery Beat.
    """
    from .models import ReportExport

    now = timezone.now()
    expired = ReportExport.objects.filter(expires_at__lt=now)
    count = expired.count()

    for export in expired:
        if export.file:
            export.file.delete(save=False)

    expired.delete()

    logger.info("Cleaned up %d expired report exports.", count)
